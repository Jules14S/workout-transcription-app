from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from google.cloud import vision
import pandas as pd
import os
import json
from io import BytesIO
from google.oauth2 import service_account
import openpyxl
from openpyxl import styles

# Path to the secret file mounted by Render
credentials_path = '/etc/secrets/google_credentials.json'

# Use the credentials to authenticate with Google Cloud Vision API
credentials = service_account.Credentials.from_service_account_file(credentials_path)
client = vision.ImageAnnotatorClient(credentials=credentials)

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})  # This will enable CORS for all routes

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def extract_text_from_image(image_path):
    """Extract text from an image using Google Cloud Vision"""
    with open(image_path, 'rb') as image_file:
        content = image_file.read()

    image = vision.Image(content=content)
    response = client.text_detection(image=image)
    texts = response.text_annotations

    if len(texts) > 0:
        return texts[0].description
    return ""

def transcribe_text_to_table(text):
    """Convert extracted text into a structured table format"""
    lines = text.split('\n')
    data = []
    max_sets = 0

    for line in lines:
        line = line.replace('.', ':')  # Replace periods with colons to fix formatting issues
        if '/' in line:
            parts = line.split(':')
            if len(parts) < 2:
                continue
            sets = parts[1].strip().split('/')
            sets = [s.strip() for s in sets if s.strip().isdigit() or s.strip() == '']
            max_sets = max(max_sets, len(sets))

    for line in lines:
        line = line.replace('.', ':')  # Replace periods with colons to fix formatting issues
        if '/' in line:
            parts = line.split(':')
            if len(parts) < 2:
                continue

            exercise = parts[0].strip()
            sets = parts[1].strip().split('/')

            sets = [s.strip() for s in sets if s.strip().isdigit() or s.strip() == '']

            while len(sets) < max_sets:
                sets.append('')

            extra_info = ""
            if '(' in parts[1]:
                extra_info = parts[1].split('(')[1].split(')')[0]

            row = [exercise] + sets
            row.append(extra_info if extra_info else '')

            data.append(row)

    return data, max_sets

def create_excel(dataframes):
    """Create an Excel file from a list of DataFrames with a custom layout"""
    output = BytesIO()

    # Use openpyxl for writing Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet(title="Workout Data")
        
        # Starting row
        start_row = 1

        # Iterate through all DataFrames, append to the same sheet with spacing
        for df_tuple in dataframes:
            try:
                df, sheet_name, date_workout_type = df_tuple  # Unpack tuple
            except ValueError as e:
                print(f"Error: {e}. Expected a tuple with 3 elements (df, sheet_name, date_workout_type). Received: {df_tuple}")
                continue

            # Add the title/date line (workout title)
            title_cell = worksheet.cell(row=start_row, column=1)
            title_cell.value = date_workout_type
            title_cell.font = openpyxl.styles.Font(bold=True)
            title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(df.columns) + 1)

            # Move down one row for the table headers
            start_row += 1

            # Write the DataFrame headers manually
            for col_num, column_title in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=start_row, column=col_num)
                cell.value = column_title
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                cell.fill = openpyxl.styles.PatternFill(start_color="00A9E0", end_color="00A9E0", fill_type="solid")
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin"),
                                                     right=openpyxl.styles.Side(border_style="thin"),
                                                     top=openpyxl.styles.Side(border_style="thin"),
                                                     bottom=openpyxl.styles.Side(border_style="thin"))

            # Move down to start writing the data
            start_row += 1

            # Write the actual data rows from DataFrame
            for row_num, row_data in enumerate(df.values, start=start_row):
                for col_num, cell_value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_num, column=col_num)

                    # Convert numeric set values from string to numbers, leave other columns unchanged
                    if df.columns[col_num - 1].startswith('Set'):
                        try:
                            # Try to convert to int or float if needed
                            if isinstance(cell_value, str) and cell_value.isdigit():
                                cell_value = int(cell_value)
                            else:
                                cell_value = float(cell_value)
                        except (ValueError, TypeError):
                            pass  # Keep original value if it can't be converted

                    cell.value = cell_value
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin"),
                                                         right=openpyxl.styles.Side(border_style="thin"),
                                                         top=openpyxl.styles.Side(border_style="thin"),
                                                         bottom=openpyxl.styles.Side(border_style="thin"))

            # Adjust start_row for the next table
            start_row += len(df) + 3  # Add 3 extra rows as spacing between tables

        # Auto-adjust column width
        for col in worksheet.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue  # Skip merged cells
                if col_letter is None:
                    col_letter = cell.column_letter  # Get the column letter
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if col_letter:
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[col_letter].width = adjusted_width

    output.seek(0)
    return output

def extract_workout_title_and_date(text):
    """Extract the workout title and date from the text."""
    lines = text.split('\n')
    workout_title = ""
    workout_date = ""
    
    for line in lines:
        line_lower = line.lower().strip()
        
        # Check if the line looks like a date (contains a number and a month or starts with "date:")
        if "date" in line_lower or any(month in line_lower for month in ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]) or any(char.isdigit() for char in line_lower):
            workout_date = line.strip()
        
        # If the line doesn't look like a date or exercise and is not empty, use it as the title
        if not workout_title and not any(char.isdigit() for char in line_lower) and not "/" in line_lower:
            workout_title = line.strip()

        # Stop as soon as both are found
        if workout_title and workout_date:
            break

    # Fallback if no title or date found
    if not workout_title:
        workout_title = "Workout"
    if not workout_date:
        workout_date = "Unknown Date"

    return f"{workout_date} - {workout_title}"


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        try:
            files = request.files.getlist('files[]')
            if not files:
                print("No files received")
                return jsonify({"error": "No files received"}), 400

            dataframes = []
            for i, file in enumerate(files):
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
                print(f"Saving file: {file.filename} at {file_path}")
                file.save(file_path)
                
                # Extract text
                text = extract_text_from_image(file_path)
                print(f"Extracted text: {text}")

                # Extract workout title and date
                workout_title_and_date = extract_workout_title_and_date(text)
                print(f"Workout title and date: {workout_title_and_date}")

                # Process the extracted text into table data
                table_data, max_sets = transcribe_text_to_table(text)
                print(f"Table data: {table_data}")

                # Create a DataFrame with the correct number of sets
                df = pd.DataFrame(table_data, columns=["Exercise"] + [f"Set {j+1}" for j in range(max_sets)] + ["Extra Info"])

                # Append the DataFrame with the correct sheet name and title
                dataframes.append((df, f"Sheet{i+1}", workout_title_and_date))

            # Generate the Excel file with multiple sheets
            excel_data = create_excel(dataframes)

            return send_file(excel_data, as_attachment=True, download_name='workout.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        except Exception as e:
            print(f"Error occurred: {e}")
            return jsonify({"error": str(e)}), 500

    return jsonify({"message": "Upload your files"}), 200

if __name__ == '__main__':
    app.run()
